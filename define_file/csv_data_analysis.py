import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

max_ave = []
min_ave = []

output_data_B = []
output_data_B_mT = []
output_data_C = []
output_data_D = []
output_data_E = []
output_data_deg_C = []
output_data_deg_D = []
f_s_name = []
xxx = 0
xx = 1

def analysis(f_name,num_d,max_d,min_d,area_d,cf_d_x,cf_d_y,min_r,max_r,area_sell,flag_inversion,xxx):
    global ave_max_data,ave_min_data
    global cf_data_x,cf_data_y,max_dt,min_dt
    
    sheetB = []
    sheetB_mT = []
    sheetC = []
    sheetD = []
    sheetE = []
    sheetdeg_C = []
    sheetdeg_D = []

    flag_max = flag_min = 0
    
    cf_data_x = cf_d_x
    cf_data_y = cf_d_y

    wb_pyxl = Workbook()
    ws_pyxl = wb_pyxl.active
    
    f_s_name.append(os.path.basename(f_name))
    
    col_names = ['c{0:02d}'.format(i) for i in range(10)]
    df = pd.read_csv(f_name,encoding = 'shift_jis',names=col_names)
    for row in dataframe_to_rows(df,index=None,header=None):
        ws_pyxl.append(row)

    sell(area_sell)

    for i in range(area_d,num_d):
        x = float(ws_pyxl.cell(row=i,column=sell_num).value)
        j = float(ws_pyxl.cell(row=i,column=sell_num+1).value)
        sheetB.append(x)
        sheetC.append(j)
        if j > max_d and max_d != 0:
            flag_max = 1
        if j < min_d and min_d != 0:
            flag_min = 1
            
    if flag_inversion == 1:
        sheetC = list(map(cal_inversion,sheetC))
        flag_inversion = 0
        
    if cf_d_x > 0:
        sheetB_mT = list(map(cal_Amf,sheetB))
            
    if flag_max == 1:
        for max_d in range(len(sheetC)):
            if max_r <= sheetC[max_d] and sheetC[max_d] < max_d:
                max_ave.append(sheetC[max_d])
        ave_max_data = sum(max_ave)/len(max_ave)
        
    if flag_min == 1:
        for min_d in range(len(sheetC)):
            if min_r > sheetC[min_d] and sheetC[min_d] <= min_d:
                min_ave.append(sheetC[min_d])
        ave_min_data = sum(min_ave)/len(min_ave)
    
    max_dt = max(sheetC)
    min_dt = min(sheetC)
    
    result = flag_jud(flag_max,flag_min)
    for z in range(len(sheetC)):
        sheetD.append(sheetC[z] - result[1])

    for y in range(len(sheetC)):
        sheetE.append(sheetD[y]/result[0])
        
    if cf_d_y > 0:
        sheetdeg_C = list(map(cal_degree,sheetC))
        sheetdeg_D = list(map(cal_degree,sheetD))
        
    xxx += 1
    
    output_data_B.append(sheetB)
    output_data_C.append(sheetC)
    output_data_D.append(sheetD)
    output_data_E.append(sheetE)
    if cf_d_x > 0:
        output_data_B_mT.append(sheetB_mT)
    if cf_d_y > 0:
        output_data_deg_C.append(sheetdeg_C)
        output_data_deg_D.append(sheetdeg_D)
        
    return  sheetB,sheetB_mT,sheetC,sheetD,sheetE,sheetdeg_C,sheetdeg_D,\
            output_data_B,output_data_B_mT,output_data_C,output_data_D,output_data_E,output_data_deg_C,output_data_deg_D,\
            xxx,f_s_name

def sell(sell):
    global sell_num
    
    area_data_sell = sell
    if area_data_sell == "A":
        sell_num = 1
    elif area_data_sell == "B":
        sell_num = 2
    elif area_data_sell == "C":
        sell_num = 3
    elif area_data_sell == "D":
        sell_num = 4
    elif area_data_sell == "E":
        sell_num = 5
    elif area_data_sell == "F":
        sell_num = 6
    elif area_data_sell == "G":
        sell_num = 7
    elif area_data_sell == "H":
        sell_num = 8
    elif area_data_sell == "I":
        sell_num = 9
    elif area_data_sell == "J":
        sell_num = 10
    elif area_data_sell == "K":
        sell_num = 11
    elif area_data_sell == "L":
        sell_num = 12
    elif area_data_sell == "M":
        sell_num = 13
    elif area_data_sell == "N":
        sell_num = 14
    elif area_data_sell == "O":
        sell_num = 15
    elif area_data_sell == "P":
        sell_num = 16
    elif area_data_sell == "Q":
        sell_num = 17
    elif area_data_sell == "R":
        sell_num = 18
    elif area_data_sell == "S":
        sell_num = 19
    elif area_data_sell == "T":
        sell_num = 20
    elif area_data_sell == "U":
        sell_num = 21
    elif area_data_sell == "V":
        sell_num = 22
    elif area_data_sell == "W":
        sell_num = 23
    elif area_data_sell == "X":
        sell_num = 24
    elif area_data_sell == "Y":
        sell_num = 25
    elif area_data_sell == "Z":
        sell_num = 26
        
def cal_inversion(n):
    return n*-1

def cal_Amf(n):
    return n*cf_data_x

def flag_jud(flag_max,flag_min):
    if flag_max != 1 and flag_min != 1:
        range_data = (max_dt-min_dt)/2
        dif_data = max_dt - range_data  
    elif flag_max == 1 and flag_min != 1:
        range_data = (ave_max_data-min_dt)/2
        dif_data = ave_max_data - range_data        
    elif flag_max != 1 and flag_min == 1:
        range_data = (max_dt-ave_min_data)/2
        dif_data = max_dt - range_data
    else:
        range_data = (ave_max_data-ave_min_data)/2
        dif_data = ave_max_data - range_data
    return range_data,dif_data

def cal_degree(n):
    return n/cf_data_y