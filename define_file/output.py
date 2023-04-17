#得られたデータから演算された結果をcsvファイルで出力する制御を行っている。
import tkinter as tk
from tkinter import filedialog as fl

from openpyxl import Workbook

sub_save = None
xx = 1

def data(f_name,x,C,outputB,outputB_mT,outputD,outputE,outputdegD,cf_x,cf_y,h,v):
    global f_s_name,xxx
    global sheetC,output_data_B,output_data_B_mT,output_data_D,output_data_E,output_data_deg_D
    global cf_data_x,cf_data_y
    global hol,ver
    
    f_s_name = f_name
    xxx = x
    sheetC = C
    output_data_B = outputB
    output_data_B_mT = outputB_mT
    output_data_D = outputD
    output_data_E = outputE
    output_data_deg_D = outputdegD
    cf_data_x = cf_x
    cf_data_y = cf_y
    hol = h
    ver = v

def output():
    global xx
    
    if rad.get() == "None":
        output = 1
    else:
        output = 0

    if rad.get() == 10 or output == 1:
        I_save = NI_save = 0
        N_save = 1
        
    elif rad.get() == 20:
        N_save = NI_save = 0
        I_save = 1
        
    elif rad.get() == 30:
        I_save = N_save = 0
        NI_save = 1

    wb_data = Workbook()
    ws_data = wb_data['Sheet']
    
    if N_save == 1:
        if cf_data_x == 0:
            if xxx > 1:
                for i in range(xxx):
                    for j in range(len(sheetC)):
                        ws_data.cell(row=1,column=i+xx).value = f_s_name[i]
                        ws_data.cell(row=2, column=i+xx).value = 'V'
                        ws_data.cell(row=2, column=i+(xx+1)).value = ver
                        ws_data.cell(row=j+3, column=i+xx,value = output_data_B[i][j])
                        ws_data.cell(row=j+3, column=i+(xx+1),value = output_data_E[i][j])
                    xx += 1

            else:
                for z in range(len(sheetC)):
                    ws_data.cell(row=1,column=1).value = f_s_name[0]
                    ws_data.cell(row=2, column=1).value = 'V'
                    ws_data.cell(row=2, column=2).value = ver
                    ws_data.cell(row=z+3, column=1,value = output_data_B[0][z])
                    ws_data.cell(row=z+3, column=2,value = output_data_E[0][z])
    
        elif cf_data_x > 0:
            if xxx > 1:
                for i in range(xxx):
                    for j in range(len(sheetC)):
                        ws_data.cell(row=1,column=i+xx).value = f_s_name[i]
                        ws_data.cell(row=2, column=i+xx).value = hol
                        ws_data.cell(row=2, column=i+(xx+1)).value = ver
                        ws_data.cell(row=j+3, column=i+xx,value = output_data_B_mT[i][j])
                        ws_data.cell(row=j+3, column=i+(xx+1),value = output_data_E[i][j])
                    xx += 1

            else:
                for z in range(len(sheetC)):
                    ws_data.cell(row=1,column=1).value = f_s_name[0]
                    ws_data.cell(row=2, column=1).value = hol
                    ws_data.cell(row=2, column=2).value = ver
                    ws_data.cell(row=z+3, column=1,value = output_data_B_mT[0][z])
                    ws_data.cell(row=z+3, column=2,value = output_data_E[0][z])

    elif I_save == 1:
        if cf_data_x == 0 and cf_data_y == 0:
            if xxx > 1:
                for i in range(xxx):
                    for j in range(len(sheetC)):
                        ws_data.cell(row=1,column=i+xx).value = f_s_name[i]
                        ws_data.cell(row=2, column=i+xx).value = 'V'
                        ws_data.cell(row=2, column=i+(xx+1)).value = 'V'
                        ws_data.cell(row=j+3, column=i+xx,value = output_data_B[i][j])
                        ws_data.cell(row=j+3, column=i+(xx+1),value = output_data_D[i][j])
                    xx += 1

            else:
                for z in range(len(sheetC)):
                    ws_data.cell(row=1,column=1).value = f_s_name[0]
                    ws_data.cell(row=2, column=1).value = 'V'
                    ws_data.cell(row=2, column=2).value = 'V'
                    ws_data.cell(row=z+3, column=1,value = output_data_B[0][z])
                    ws_data.cell(row=z+3, column=2,value = output_data_D[0][z])
                    
        elif cf_data_x > 0:
            if xxx > 1:
                for i in range(xxx):
                    for j in range(len(sheetC)):
                        ws_data.cell(row=1,column=i+xx).value = f_s_name[i]
                        ws_data.cell(row=2, column=i+xx).value = hol
                        ws_data.cell(row=2, column=i+(xx+1)).value = ver
                        ws_data.cell(row=j+3, column=i+xx,value = output_data_B_mT[i][j])
                        if cf_data_y > 0.00:
                            ws_data.cell(row=j+3, column=i+(xx+1),value = output_data_deg_D[i][j])
                        else:
                            ws_data.cell(row=j+3, column=i+(xx+1),value = output_data_D[i][j])
                    xx += 1

            else:
                for z in range(len(sheetC)):
                    ws_data.cell(row=1,column=1).value = f_s_name[0]
                    ws_data.cell(row=2, column=1).value = hol
                    ws_data.cell(row=2, column=2).value = ver
                    ws_data.cell(row=z+3, column=1,value = output_data_B_mT[0][z])
                    if cf_data_y > 0.00:
                        ws_data.cell(row=z+3, column=2,value = output_data_deg_D[0][z])
                    else:
                        ws_data.cell(row=z+3, column=2,value = output_data_D[0][z])
                        
        elif cf_data_y > 0.00:
            if xxx > 1:
                for i in range(xxx):
                    for j in range(len(sheetC)):
                        ws_data.cell(row=1,column=i+xx).value = f_s_name[i]
                        ws_data.cell(row=2, column=i+xx).value = hol
                        ws_data.cell(row=2, column=i+(xx+1)).value = ver
                        if cf_data_x > 0:
                            ws_data.cell(row=j+3, column=i+xx,value = output_data_B_mT[i][j])
                        else:
                            ws_data.cell(row=j+3, column=i+xx,value = output_data_B[i][j])
                        ws_data.cell(row=j+3, column=i+(xx+1),value = output_data_deg_D[i][j])
                        ws_data.cell(row=j+3, column=i+(xx+1),value = output_data_D[i][j])
                    xx += 1

            else:
                for z in range(len(sheetC)):
                    ws_data.cell(row=1,column=1).value = f_s_name[0]
                    ws_data.cell(row=2, column=1).value = hol
                    ws_data.cell(row=2, column=2).value = ver
                    if cf_data_x > 0:
                        ws_data.cell(row=z+3, column=1,value = output_data_B_mT[0][z])
                    else:
                        ws_data.cell(row=z+3, column=1,value = output_data_B_mT[0][z])
                    ws_data.cell(row=z+3, column=2,value = output_data_deg_D[0][z])
                    ws_data.cell(row=z+3, column=2,value = output_data_D[0][z])
                    
    elif NI_save == 1:
        wb_data_I = Workbook()
        ws_data_I = wb_data_I['Sheet']
        if cf_data_x == 0:
            if xxx > 1:
                for i in range(xxx):
                    for j in range(len(sheetC)):
                        ws_data.cell(row=1,column=i+xx).value = f_s_name[i]
                        ws_data.cell(row=2, column=i+xx).value = 'V'
                        ws_data.cell(row=2, column=i+(xx+1)).value = ver
                        ws_data.cell(row=j+3, column=i+xx,value = output_data_B[i][j])
                        ws_data.cell(row=j+3, column=i+(xx+1),value = output_data_E[i][j])
                        
                        ws_data_I.cell(row=1,column=i+xx).value = f_s_name[i]
                        ws_data_I.cell(row=2, column=i+xx).value = 'V'
                        ws_data_I.cell(row=2, column=i+(xx+1)).value = 'V'
                        ws_data_I.cell(row=j+3, column=i+xx,value = output_data_B[i][j])
                        ws_data_I.cell(row=j+3, column=i+(xx+1),value = output_data_D[i][j])
                    xx += 1

            else:
                for z in range(len(sheetC)):
                    ws_data.cell(row=1,column=1).value = f_s_name[0]
                    ws_data.cell(row=2, column=1).value = 'V'
                    ws_data.cell(row=2, column=2).value = ver
                    ws_data.cell(row=z+3, column=1,value = output_data_B[0][z])
                    ws_data.cell(row=z+3, column=2,value = output_data_E[0][z])
                    
                    ws_data_I.cell(row=1,column=1).value = f_s_name[0]
                    ws_data_I.cell(row=2, column=1).value = 'V'
                    ws_data_I.cell(row=2, column=2).value = 'V'
                    ws_data_I.cell(row=z+3, column=1,value = output_data_B[0][z])
                    ws_data_I.cell(row=z+3, column=2,value = output_data_D[0][z])
    
        elif cf_data_x > 0:
            if xxx > 1:
                for i in range(xxx):
                    for j in range(len(sheetC)):
                        ws_data.cell(row=1,column=i+xx).value = f_s_name[i]
                        ws_data.cell(row=2, column=i+xx).value = hol
                        ws_data.cell(row=2, column=i+(xx+1)).value = ver
                        ws_data.cell(row=j+3, column=i+xx,value = output_data_B_mT[i][j])
                        ws_data.cell(row=j+3, column=i+(xx+1),value = output_data_E[i][j])
                    xx += 1

            else:
                for z in range(len(sheetC)):
                    ws_data.cell(row=1,column=1).value = f_s_name[0]
                    ws_data.cell(row=2, column=1).value = hol
                    ws_data.cell(row=2, column=2).value = ver
                    ws_data.cell(row=z+3, column=1,value = output_data_B_mT[0][z])
                    ws_data.cell(row=z+3, column=2,value = output_data_E[0][z])
    
    filename = fl.asksaveasfilename(
    title = "名前を付けて保存",
    filetypes = [("CSV",".csv"),("Excel",".xlsx") ], # ファイルフィルタ
    initialdir = "./", # 自分自身のディレクトリ
    defaultextension = "csv"
    )
    
    target = '.'
    idx = filename.find(target)
    r1 = filename[:idx]
    r2 = filename[idx:]
    
    if N_save == 1 or I_save == 1:
        wb_data.save(filename)
        
    elif NI_save == 1:
        wb_data.save(r1 + "_1" + r2)
        wb_data_I.save(r1 + "_2" + r2)

def output_det(event=None):
    global sub_save,rad
    rad = {}
    
    if sub_save == None or not sub_save.winfo_exists():
        sub_save = tk.Toplevel()
        sub_save.geometry("370x240")
        sub_save.title("保存データの指定")
        sub_save.configure(bg='pale green')
        sub_save.resizable(0,0)
        
        label = tk.Label(sub_save,text='保存したいデータを選択してください',font=("明朝体", "15", "bold"))
        label.pack(anchor='center',pady=5)
        
        frame = tk.Frame(
            sub_save,relief=tk.RIDGE,width=300,height=190,bd=3).place(x=6,y=40)
        
        rad = tk.IntVar()
        rad.set(10)
        
        
        rad_save_1 = tk.Radiobutton(sub_save,value=10,variable=rad,text="規格化データのみ",font=("明朝体", "15", "bold"))
        rad_save_1.place(x=10,y=60)
        
        rad_save_2 = tk.Radiobutton(sub_save,value=20,variable=rad,text="初期データのみ",font=("明朝体", "15", "bold"))
        rad_save_2.place(x=10,y=110)
        
        rad_save_3 = tk.Radiobutton(sub_save,value=30,variable=rad,text="規格化データと初期データ",font=("明朝体", "15", "bold"))
        rad_save_3.place(x=10,y=160)
        
        save_select_button = tk.Button(
            sub_save,text='保存',font=("明朝体","15","bold"),foreground='black',command=output).place(x=312,y=190,width=50)
